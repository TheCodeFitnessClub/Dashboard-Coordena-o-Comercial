"""
The Code — Closum Processor
============================
Automatiza o cálculo do tempo de resposta a leads novas (Closum → 1º contacto)
e atualiza a aba 02_Leads_Closum de The_Code_Workflow.xlsx.

USO BÁSICO
----------
    python3 closum_processor.py path/para/closum_export.csv

USO COM WATCH (corre a cada 5 min monitorizando uma pasta)
----------------------------------------------------------
    python3 closum_processor.py --watch path/para/pasta_de_exports

ALTERAR MAPEAMENTO DE COLUNAS DO CLOSUM
---------------------------------------
Edita o dicionário FIELD_MAP abaixo (linhas 35-50) para corresponder aos
nomes exatos das colunas no teu export do Closum.

REQUISITOS
----------
    pip install pandas openpyxl

OUTPUT
------
- Atualiza The_Code_Workflow.xlsx (aba 02_Leads_Closum)
- Imprime relatório no terminal com:
    · Total de leads processadas
    · Tempo médio de resposta
    · Leads em violação SLA (>60min)
    · Quebra por comercial
- Cria closum_log.txt com histórico de execuções
"""

from __future__ import annotations
import argparse
import sys
import time
import csv
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook

# ============================================================
# CONFIG — adaptar aos cabeçalhos reais do Closum
# ============================================================
# Esquerda = nome no nosso workbook | Direita = nome no CSV do Closum
FIELD_MAP = {
    "Lead ID":                       ["lead_id", "id", "Lead ID"],
    "Nome":                          ["name", "nome", "Nome", "full_name", "lead_name"],
    "Telefone":                      ["phone", "telefone", "Telefone", "phone_number"],
    "Origem":                        ["source", "origem", "Origem", "lead_source", "utm_source"],
    "Entrada Closum (data/hora)":    ["created_at", "data_criacao", "creation_date", "Created At", "Data Criação"],
    "Comercial":                     ["assigned_to", "comercial", "Comercial", "owner", "Assigned To"],
    "Primeiro contacto (data/hora)": ["first_contact_at", "data_primeiro_contacto", "first_contacted_at", "First Contact"],
    "Canal":                         ["first_contact_channel", "canal", "Canal", "contact_channel"],
    "Resultado":                     ["status", "estado", "Resultado", "result"],
    "Notas":                         ["notes", "notas", "Notas", "comments"],
}

# Mapeamento de comerciais (caso o Closum guarde emails ou IDs em vez de nomes)
COMERCIAL_ALIAS = {
    # "user_id_123": "Nome Comercial CU",
    # "comercial.cu@thecode.pt": "Nome Comercial CU",
}

WORKBOOK_PATH = Path(__file__).parent / "The_Code_Workflow.xlsx"
LEADS_SHEET = "02_Leads_Closum"
LOG_FILE = Path(__file__).parent / "closum_log.txt"

SLA_MINUTES = 60  # Manual: primeiro contacto até 1h


# ============================================================
# Utilidades
# ============================================================
def log(msg: str):
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line)
    with LOG_FILE.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def find_column(headers: list[str], candidates: list[str]) -> str | None:
    """Encontra a primeira coluna que bate com a lista de aliases."""
    headers_lower = {h.lower().strip(): h for h in headers}
    for c in candidates:
        if c.lower().strip() in headers_lower:
            return headers_lower[c.lower().strip()]
    return None


def parse_dt(value) -> datetime | None:
    """Tenta vários formatos de datetime comuns no Closum."""
    if value is None or value == "" or str(value).strip().lower() in ("nan", "none", "null"):
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S.%fZ",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def read_csv(path: Path) -> list[dict]:
    """Lê CSV detetando delimitador e encoding."""
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
    last_err = None
    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.DictReader(f, dialect=dialect)
                return list(reader)
        except UnicodeDecodeError as e:
            last_err = e
            continue
    raise RuntimeError(f"Não consegui ler o CSV com encodings comuns: {last_err}")


# ============================================================
# Núcleo
# ============================================================
def process_leads(csv_path: Path) -> list[dict]:
    """Processa o CSV do Closum e devolve lista de dicts pronta para o Excel."""
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV não encontrado: {csv_path}")

    rows = read_csv(csv_path)
    if not rows:
        log(f"⚠ CSV vazio: {csv_path}")
        return []

    headers = list(rows[0].keys())
    col_map = {target: find_column(headers, candidates) for target, candidates in FIELD_MAP.items()}

    missing = [t for t, c in col_map.items() if c is None and t in (
        "Nome", "Entrada Closum (data/hora)"
    )]
    if missing:
        raise RuntimeError(
            f"Colunas obrigatórias não encontradas no CSV: {missing}\n"
            f"Cabeçalhos disponíveis: {headers}\n"
            f"Edita FIELD_MAP em closum_processor.py para mapear corretamente."
        )

    processed = []
    for r in rows:
        lead = {}
        for target, source in col_map.items():
            lead[target] = r.get(source, "") if source else ""

        # Resolver alias de comercial
        com = lead.get("Comercial", "")
        if com in COMERCIAL_ALIAS:
            lead["Comercial"] = COMERCIAL_ALIAS[com]

        # Calcular tempo de resposta
        dt_in = parse_dt(lead.get("Entrada Closum (data/hora)"))
        dt_first = parse_dt(lead.get("Primeiro contacto (data/hora)"))

        if dt_in and dt_first:
            delta_min = (dt_first - dt_in).total_seconds() / 60.0
            lead["_tempo_min"] = max(0, round(delta_min))
            lead["_sla"] = "OK" if delta_min <= SLA_MINUTES else "VIOLADO"
        else:
            lead["_tempo_min"] = None
            lead["_sla"] = "Sem contacto" if dt_in and not dt_first else None

        lead["_dt_in"] = dt_in
        lead["_dt_first"] = dt_first
        processed.append(lead)

    return processed


def update_workbook(leads: list[dict], wb_path: Path = WORKBOOK_PATH):
    """Atualiza a aba 02_Leads_Closum preservando fórmulas e formatação."""
    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook não encontrado: {wb_path}")

    wb = load_workbook(wb_path)
    if LEADS_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Aba '{LEADS_SHEET}' não existe em {wb_path}")
    ws = wb[LEADS_SHEET]

    # Limpar linhas de dados (manter título e cabeçalho — linhas 1 e 2)
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=3, max_row=max_row, max_col=12):
        for cell in row:
            # Não apagar fórmulas das colunas I (tempo) e J (SLA)
            if cell.column in (9, 10):
                continue
            cell.value = None

    # Escrever novos dados
    for idx, lead in enumerate(leads, start=3):
        ws.cell(row=idx, column=1, value=lead.get("Lead ID", ""))
        ws.cell(row=idx, column=2, value=lead.get("Nome", ""))
        ws.cell(row=idx, column=3, value=lead.get("Telefone", ""))
        ws.cell(row=idx, column=4, value=lead.get("Origem", ""))
        if lead.get("_dt_in"):
            ws.cell(row=idx, column=5, value=lead["_dt_in"])
        ws.cell(row=idx, column=6, value=lead.get("Comercial", ""))
        if lead.get("_dt_first"):
            ws.cell(row=idx, column=7, value=lead["_dt_first"])
        ws.cell(row=idx, column=8, value=lead.get("Canal", ""))
        # I (tempo) e J (SLA) — fórmulas já existem
        ws.cell(row=idx, column=11, value=lead.get("Resultado", ""))
        ws.cell(row=idx, column=12, value=lead.get("Notas", ""))

    wb.save(wb_path)
    log(f"✓ Workbook atualizado: {len(leads)} leads escritas em {wb_path.name}")


def report(leads: list[dict]):
    """Imprime resumo no terminal."""
    if not leads:
        log("Nenhuma lead para reportar.")
        return

    total = len(leads)
    com_contacto = [l for l in leads if l["_tempo_min"] is not None]
    violados = [l for l in com_contacto if l["_sla"] == "VIOLADO"]
    sem_contacto = [l for l in leads if l["_sla"] == "Sem contacto"]

    tempos = [l["_tempo_min"] for l in com_contacto]
    avg = round(sum(tempos) / len(tempos), 1) if tempos else 0

    print()
    print("=" * 60)
    print("RELATÓRIO CLOSUM — TEMPO DE RESPOSTA")
    print("=" * 60)
    print(f"Total leads processadas:    {total}")
    print(f"Com primeiro contacto:      {len(com_contacto)}")
    print(f"Sem contacto registado:     {len(sem_contacto)}")
    print(f"SLA cumprido (≤{SLA_MINUTES}min):     {len(com_contacto) - len(violados)}")
    print(f"SLA violado (>{SLA_MINUTES}min):      {len(violados)}")
    print(f"Tempo médio de resposta:    {avg} min")
    print()

    # Quebra por comercial
    by_com = {}
    for l in leads:
        c = l.get("Comercial") or "—"
        by_com.setdefault(c, []).append(l)

    print("POR COMERCIAL")
    print("-" * 60)
    print(f"{'Comercial':<28}{'Leads':>8}{'T.médio':>10}{'Violações':>14}")
    for c, ls in sorted(by_com.items()):
        ts = [l["_tempo_min"] for l in ls if l["_tempo_min"] is not None]
        avg_c = round(sum(ts) / len(ts), 1) if ts else 0
        viol_c = sum(1 for l in ls if l["_sla"] == "VIOLADO")
        print(f"{c[:28]:<28}{len(ls):>8}{avg_c:>10}{viol_c:>14}")
    print()

    if violados:
        print("LEADS COM SLA VIOLADO (TOP 10)")
        print("-" * 60)
        for l in sorted(violados, key=lambda x: -x["_tempo_min"])[:10]:
            print(f"  · {l.get('Nome','—')[:30]:<30} {l['_tempo_min']:>5} min  ({l.get('Comercial','—')})")

    if sem_contacto:
        print()
        print(f"⚠ {len(sem_contacto)} LEADS AINDA POR CONTACTAR — agir já.")


def run_once(csv_path: Path):
    log(f"A processar {csv_path.name}…")
    leads = process_leads(csv_path)
    update_workbook(leads)
    report(leads)


def watch(folder: Path, interval_seconds: int = 300):
    log(f"Watcher iniciado em {folder} (intervalo: {interval_seconds}s)")
    seen = {}
    while True:
        try:
            csvs = sorted(folder.glob("*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
            if csvs:
                latest = csvs[0]
                mtime = latest.stat().st_mtime
                if seen.get(str(latest)) != mtime:
                    seen[str(latest)] = mtime
                    log(f"Novo export detetado: {latest.name}")
                    run_once(latest)
            time.sleep(interval_seconds)
        except KeyboardInterrupt:
            log("Watcher interrompido pelo utilizador.")
            break
        except Exception as e:
            log(f"⚠ Erro no watcher: {e}")
            time.sleep(interval_seconds)


def main():
    parser = argparse.ArgumentParser(description="Processador de exports Closum para o The Code Workflow")
    parser.add_argument("path", help="Caminho para o CSV (modo único) ou pasta (modo --watch)")
    parser.add_argument("--watch", action="store_true", help="Monitorizar pasta e reprocessar quando aparecem novos CSV")
    parser.add_argument("--interval", type=int, default=300, help="Intervalo do watcher em segundos (default: 300)")
    args = parser.parse_args()

    p = Path(args.path)
    if args.watch:
        if not p.is_dir():
            print(f"Erro: {p} não é uma pasta.")
            sys.exit(1)
        watch(p, args.interval)
    else:
        if not p.is_file():
            print(f"Erro: {p} não é um ficheiro.")
            sys.exit(1)
        run_once(p)


if __name__ == "__main__":
    main()
